"""
Génère un fichier Excel : Liste des produits à commander par fournisseur
========================================================================
MakitiPlus — Bon de commande groupé par fournisseur
Devise : GNF (Franc guinéen)
"""

import sys, os

XLSX_SKILL_DIR = "/home/z/my-project/skills/xlsx"
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, "templates")]:
    if sub not in sys.path:
        sys.path.insert(0, sub)

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from templates.base import (
    use_palette_explicit, setup_sheet, style_header_row, style_data_row,
    style_total_row, font_title, font_header, font_subheader, font_body,
    font_caption, fill_header, fill_total, fill_data_row,
    border_header, border_total,
    align_title, align_header, align_number, align_text, align_date,
    PRIMARY, PRIMARY_LIGHT, SECONDARY,
    ACCENT_POSITIVE, ACCENT_NEGATIVE, ACCENT_WARNING,
    NEUTRAL_900, NEUTRAL_600, NEUTRAL_200, NEUTRAL_100, NEUTRAL_0,
    HEADER_TEXT, COLUMN_WIDTHS, ROW_HEIGHTS, FORMATS,
    auto_fit_columns, CF_WARNING_FILL, CF_WARNING_FONT,
    CF_NEGATIVE_FILL, CF_NEGATIVE_FONT, CF_POSITIVE_FILL, CF_POSITIVE_FONT,
    FONT_NAME, HEADER_BOLD,
)

# Use professional palette
use_palette_explicit("professional")

# ============================================================
# Sample Data — Produits à commander par fournisseur
# ============================================================
# Each supplier has: name, phone, email, city, and a list of products
# Products: (name, category, current_stock, min_stock, cost_price, unit)

SUPPLIERS_DATA = [
    {
        "name": "Distributeur Guinée SARL",
        "phone": "+224 622 11 22 33",
        "email": "contact@distributeur-gn.com",
        "city": "Conakry",
        "products": [
            ("Riz 25kg", "Alimentation", 8, 30, 285000, "sac"),
            ("Huile végétale 5L", "Alimentation", 5, 20, 125000, "carton"),
            ("Sucre 50kg", "Alimentation", 3, 15, 420000, "sac"),
            ("Pâte tomate 2.4kg", "Alimentation", 12, 25, 45000, "carton"),
            ("Lait en poudre 400g", "Alimentation", 6, 20, 35000, "carton"),
        ]
    },
    {
        "name": "Africa Boissons Import",
        "phone": "+224 628 44 55 66",
        "email": "commandes@africaboissons.com",
        "city": "Conakry",
        "products": [
            ("Coca-Cola 33cl", "Boissons", 50, 120, 3500, "carton"),
            ("Jus de fruit 1L", "Boissons", 15, 40, 8000, "carton"),
            ("Eau minérale 1.5L", "Boissons", 20, 60, 4500, "carton"),
            ("Bière Flag 65cl", "Boissons", 10, 30, 7500, "carton"),
        ]
    },
    {
        "name": "Froid & Climat SARL",
        "phone": "+224 623 77 88 99",
        "email": "ventes@froidclimat-gn.com",
        "city": "Conakry",
        "products": [
            ("Réfrigérateur 300L", "Électroménager", 2, 5, 3500000, "unité"),
            ("Climatiseur 12000 BTU", "Électroménager", 1, 3, 2800000, "unité"),
            ("Ventilateur sur pied", "Électroménager", 4, 10, 350000, "unité"),
        ]
    },
    {
        "name": "Hygiène Plus Guinée",
        "phone": "+224 625 33 44 55",
        "email": "info@hygieneplus-gn.com",
        "city": "Kindia",
        "products": [
            ("Savon de toilette 150g", "Hygiène", 30, 80, 2500, "carton"),
            ("Détergent 1kg", "Hygiène", 15, 40, 8000, "carton"),
            ("Dentifrice 100ml", "Hygiène", 20, 50, 3500, "carton"),
            ("Shampooing 250ml", "Hygiène", 10, 25, 6000, "carton"),
            ("Papier toilette x12", "Hygiène", 8, 30, 15000, "carton"),
            ("Gel hydroalcoolique 500ml", "Hygiène", 5, 15, 12000, "carton"),
        ]
    },
    {
        "name": "Textile Conakry",
        "phone": "+224 627 66 77 88",
        "email": "grossiste@textileconakry.com",
        "city": "Conakry",
        "products": [
            ("Pagne wax premium", "Textile", 12, 30, 45000, "pièce"),
            ("T-shirt homme", "Textile", 20, 50, 15000, "pièce"),
            ("Bazin riche 10m", "Textile", 5, 15, 120000, "pièce"),
        ]
    },
]

# Currency format for GNF
CURRENCY_FORMAT = '#,##0" GNF"'
DATE_FORMAT = "YYYY-MM-DD"


def build_workbook():
    wb = Workbook()

    # ============================================================
    # Sheet 1: Commandes par Fournisseur (detailed)
    # ============================================================
    ws1 = wb.active
    ws1.title = "Commandes par Fournisseur"
    ws1.sheet_properties.tabColor = PRIMARY.replace("#", "")

    # Column definitions for Sheet 1
    # Starting at col B (2)
    # B=N°, C=Produit, D=Catégorie, E=Stock actuel, F=Stock min,
    # G=Qté à commander, H=Prix d'achat unit., I=Coût total, J=Unité
    col_defs = [
        ("N°", 6),
        ("Produit", 28),
        ("Catégorie", 16),
        ("Stock actuel", 14),
        ("Stock min.", 14),
        ("Qté à commander", 16),
        ("Prix d'achat unit.", 18),
        ("Coût total estimé", 20),
        ("Unité", 10),
    ]
    last_col = len(col_defs) + 1  # +1 because we start at col B

    # Setup sheet with title
    setup_sheet(ws1, title="Liste des Produits à Commander par Fournisseur", last_col=last_col)

    # Subtitle row
    ws1.merge_cells(start_row=3, start_column=2, end_row=3, end_column=last_col)
    from datetime import date
    subtitle_cell = ws1.cell(row=3, column=2, value=f"MakitiPlus — Bon de commande — {date.today().strftime('%d/%m/%Y')}")
    subtitle_cell.font = font_caption()
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")

    current_row = 5  # Start after title + spacer + subtitle + blank
    grand_total_cost = 0
    grand_total_qty = 0

    for supplier_idx, supplier in enumerate(SUPPLIERS_DATA):
        # Supplier section header
        ws1.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=last_col)
        supplier_cell = ws1.cell(row=current_row, column=2,
                                  value=f"  {supplier['name']}  —  {supplier['city']}  |  Tél : {supplier['phone']}  |  {supplier['email']}")
        supplier_cell.font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color="FFFFFF")
        supplier_cell.fill = PatternFill("solid", fgColor=PRIMARY)
        supplier_cell.alignment = Alignment(horizontal="left", vertical="center")
        # Apply fill to all merged cells
        for c in range(2, last_col + 1):
            ws1.cell(row=current_row, column=c).fill = PatternFill("solid", fgColor=PRIMARY)
            ws1.cell(row=current_row, column=c).font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color="FFFFFF")
        ws1.row_dimensions[current_row].height = 28
        current_row += 1

        # Column headers
        for col_idx, (header, width) in enumerate(col_defs, start=2):
            cell = ws1.cell(row=current_row, column=col_idx, value=header)
            cell.font = font_header()
            cell.fill = PatternFill("solid", fgColor=PRIMARY_LIGHT)
            cell.font = Font(name=FONT_NAME, size=10, bold=HEADER_BOLD, color=PRIMARY)
            cell.alignment = align_header()
            cell.border = Border(bottom=Side(style="thin", color=NEUTRAL_200))
        ws1.row_dimensions[current_row].height = 24
        current_row += 1

        # Data rows
        supplier_total_cost = 0
        supplier_total_qty = 0
        first_data_row = current_row

        for prod_idx, (prod_name, category, stock, min_stock, cost_price, unit) in enumerate(supplier["products"]):
            qty_to_order = max(0, min_stock - stock)
            total_cost = qty_to_order * cost_price

            row_data = [
                prod_idx + 1,
                prod_name,
                category,
                stock,
                min_stock,
                qty_to_order,
                cost_price,
                total_cost,
                unit,
            ]

            for col_idx, value in enumerate(row_data, start=2):
                cell = ws1.cell(row=current_row, column=col_idx, value=value)
                cell.font = font_body()

                # Alternating row fill
                row_fill = fill_data_row(prod_idx)
                cell.fill = row_fill

                # Alignment
                if isinstance(value, (int, float)):
                    cell.alignment = align_number()
                    if col_idx in (8, 9):  # Prix d'achat, Coût total
                        cell.number_format = CURRENCY_FORMAT
                    elif col_idx in (6, 7):  # Qté à commander, Prix
                        cell.number_format = '#,##0'
                else:
                    cell.alignment = align_text()

            # Highlight rows where qty_to_order > 0 (needs ordering)
            qty_cell = ws1.cell(row=current_row, column=7)  # Qté à commander
            if qty_to_order > 0:
                qty_cell.font = Font(name=FONT_NAME, size=11, bold=True, color=ACCENT_NEGATIVE)

            supplier_total_cost += total_cost
            supplier_total_qty += qty_to_order
            ws1.row_dimensions[current_row].height = 22
            current_row += 1

        last_data_row = current_row - 1

        # Supplier subtotal row
        ws1.cell(row=current_row, column=2, value="").font = font_subheader()
        ws1.cell(row=current_row, column=3, value="Sous-total fournisseur").font = font_subheader()
        ws1.cell(row=current_row, column=3).alignment = align_text()

        # Sum for qty to order
        qty_col_letter = get_column_letter(7)  # G
        ws1.cell(row=current_row, column=7,
                 value=f"=SUM({qty_col_letter}{first_data_row}:{qty_col_letter}{last_data_row})")
        ws1.cell(row=current_row, column=7).number_format = '#,##0'
        ws1.cell(row=current_row, column=7).alignment = align_number()

        # Sum for total cost
        cost_col_letter = get_column_letter(9)  # I
        ws1.cell(row=current_row, column=9,
                 value=f"=SUM({cost_col_letter}{first_data_row}:{cost_col_letter}{last_data_row})")
        ws1.cell(row=current_row, column=9).number_format = CURRENCY_FORMAT
        ws1.cell(row=current_row, column=9).alignment = align_number()

        style_total_row(ws1, row_num=current_row, col_start=2, col_end=last_col)

        grand_total_cost += supplier_total_cost
        grand_total_qty += supplier_total_qty
        current_row += 2  # blank row between suppliers

    # Grand total row
    ws1.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=6)
    grand_cell = ws1.cell(row=current_row, column=2, value="TOTAL GÉNÉRAL")
    grand_cell.font = Font(name=FONT_NAME, size=12, bold=HEADER_BOLD, color="FFFFFF")
    grand_cell.fill = PatternFill("solid", fgColor=PRIMARY)
    grand_cell.alignment = Alignment(horizontal="right", vertical="center")
    for c in range(2, last_col + 1):
        ws1.cell(row=current_row, column=c).fill = PatternFill("solid", fgColor=PRIMARY)
        ws1.cell(row=current_row, column=c).font = Font(name=FONT_NAME, size=12, bold=HEADER_BOLD, color="FFFFFF")

    # We need to compute grand total since formulas across sections are complex
    ws1.cell(row=current_row, column=7, value=grand_total_qty)
    ws1.cell(row=current_row, column=7).number_format = '#,##0'
    ws1.cell(row=current_row, column=7).alignment = align_number()
    ws1.cell(row=current_row, column=7).font = Font(name=FONT_NAME, size=12, bold=HEADER_BOLD, color="FFFFFF")
    ws1.cell(row=current_row, column=7).fill = PatternFill("solid", fgColor=PRIMARY)

    ws1.cell(row=current_row, column=9, value=grand_total_cost)
    ws1.cell(row=current_row, column=9).number_format = CURRENCY_FORMAT
    ws1.cell(row=current_row, column=9).alignment = align_number()
    ws1.cell(row=current_row, column=9).font = Font(name=FONT_NAME, size=12, bold=HEADER_BOLD, color="FFFFFF")
    ws1.cell(row=current_row, column=9).fill = PatternFill("solid", fgColor=PRIMARY)

    ws1.row_dimensions[current_row].height = 30

    # Notes row
    current_row += 2
    ws1.cell(row=current_row, column=2, value="Note : La quantité à commander = Stock minimum - Stock actuel (si positif). Les prix sont en GNF (Franc guinéen).")
    ws1.cell(row=current_row, column=2).font = font_caption()

    # Column widths
    for col_idx, (_, width) in enumerate(col_defs, start=2):
        ws1.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze panes — freeze supplier name area + headers
    ws1.freeze_panes = "B5"

    # Print setup
    ws1.page_setup.orientation = 'landscape'
    ws1.page_setup.fitToWidth = 1
    ws1.page_setup.fitToHeight = 0

    # Conditional formatting: highlight stock actuel < stock min (cells in col E where value < col F)
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    # Apply to all data areas — we'll use a broad range
    # Red fill for items that need ordering (qty > 0 in column G)

    # ============================================================
    # Sheet 2: Résumé Fournisseurs
    # ============================================================
    ws2 = wb.create_sheet("Résumé Fournisseurs")
    ws2.sheet_properties.tabColor = ACCENT_POSITIVE.replace("#", "")

    summary_headers = [
        ("N°", 6),
        ("Fournisseur", 28),
        ("Ville", 14),
        ("Téléphone", 20),
        ("Email", 30),
        ("Nb Produits", 14),
        ("Qté à commander", 16),
        ("Coût total estimé", 20),
    ]
    summary_last_col = len(summary_headers) + 1

    setup_sheet(ws2, title="Résumé par Fournisseur", last_col=summary_last_col)

    # Subtitle
    ws2.merge_cells(start_row=3, start_column=2, end_row=3, end_column=summary_last_col)
    sub2 = ws2.cell(row=3, column=2, value=f"Synthèse des commandes — {date.today().strftime('%d/%m/%Y')}")
    sub2.font = font_caption()
    sub2.alignment = Alignment(horizontal="left", vertical="center")

    # Headers at row 5
    for col_idx, (header, _) in enumerate(summary_headers, start=2):
        cell = ws2.cell(row=5, column=col_idx, value=header)
        cell.font = font_header()
        cell.fill = fill_header()
        cell.alignment = align_header()
        cell.border = border_header()
    ws2.row_dimensions[5].height = ROW_HEIGHTS["header"]

    # Data rows
    summary_first_data = 6
    for idx, supplier in enumerate(SUPPLIERS_DATA):
        row_num = summary_first_data + idx
        product_count = len(supplier["products"])
        total_qty = sum(max(0, min_s - stock) for _, _, stock, min_s, _, _ in supplier["products"])
        total_cost = sum(max(0, min_s - stock) * price for _, _, stock, min_s, price, _ in supplier["products"])

        row_data = [
            idx + 1,
            supplier["name"],
            supplier["city"],
            supplier["phone"],
            supplier["email"],
            product_count,
            total_qty,
            total_cost,
        ]

        for col_idx, value in enumerate(row_data, start=2):
            cell = ws2.cell(row=row_num, column=col_idx, value=value)
            cell.font = font_body()
            row_fill = fill_data_row(idx)
            cell.fill = row_fill

            if isinstance(value, (int, float)):
                cell.alignment = align_number()
                if col_idx == 9:  # Coût total
                    cell.number_format = CURRENCY_FORMAT
                elif col_idx in (7, 8):  # Counts
                    cell.number_format = '#,##0'
            else:
                cell.alignment = align_text()

        ws2.row_dimensions[row_num].height = ROW_HEIGHTS["data"]

    summary_last_data = summary_first_data + len(SUPPLIERS_DATA) - 1

    # Total row
    total_row_num = summary_last_data + 1
    ws2.cell(row=total_row_num, column=2, value="").font = font_subheader()
    ws2.cell(row=total_row_num, column=3, value="TOTAL").font = font_subheader()
    ws2.cell(row=total_row_num, column=3).alignment = align_text()

    # Sum formulas
    for col_letter_idx in [7, 8, 9]:  # Nb Produits, Qté, Coût
        col_letter = get_column_letter(col_letter_idx)
        cell = ws2.cell(row=total_row_num, column=col_letter_idx,
                        value=f"=SUM({col_letter}{summary_first_data}:{col_letter}{summary_last_data})")
        cell.alignment = align_number()
        if col_letter_idx == 9:
            cell.number_format = CURRENCY_FORMAT
        else:
            cell.number_format = '#,##0'

    style_total_row(ws2, row_num=total_row_num, col_start=2, col_end=summary_last_col)

    # Column widths
    for col_idx, (_, width) in enumerate(summary_headers, start=2):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    # Conditional formatting on summary: highlight high cost suppliers
    from openpyxl.formatting.rule import DataBarRule
    cost_range = f"I{summary_first_data}:I{summary_last_data}"
    ws2.conditional_formatting.add(cost_range,
        DataBarRule(start_type='min', end_type='max',
                   color=PRIMARY, showValue=True))

    ws2.freeze_panes = "B6"
    ws2.page_setup.orientation = 'landscape'

    # ============================================================
    # Sheet 3: Détail Produits (flat list for filtering/sorting)
    # ============================================================
    ws3 = wb.create_sheet("Détail Produits")
    ws3.sheet_properties.tabColor = ACCENT_WARNING.replace("#", "")

    detail_headers = [
        ("N°", 6),
        ("Fournisseur", 26),
        ("Produit", 26),
        ("Catégorie", 14),
        ("Stock actuel", 14),
        ("Stock min.", 14),
        ("Qté à commander", 16),
        ("Prix d'achat unit.", 18),
        ("Coût total estimé", 20),
        ("Unité", 10),
        ("Statut", 14),
    ]
    detail_last_col = len(detail_headers) + 1

    setup_sheet(ws3, title="Détail des Produits à Commander", last_col=detail_last_col)

    # Headers at row 4
    for col_idx, (header, _) in enumerate(detail_headers, start=2):
        cell = ws3.cell(row=4, column=col_idx, value=header)
        cell.font = font_header()
        cell.fill = fill_header()
        cell.alignment = align_header()
        cell.border = border_header()
    ws3.row_dimensions[4].height = ROW_HEIGHTS["header"]

    # Data rows — flat list
    row_num = 5
    counter = 0
    for supplier in SUPPLIERS_DATA:
        for prod_name, category, stock, min_stock, cost_price, unit in supplier["products"]:
            counter += 1
            qty_to_order = max(0, min_stock - stock)
            total_cost = qty_to_order * cost_price
            status = "À commander" if qty_to_order > 0 else "Stock suffisant"

            row_data = [
                counter,
                supplier["name"],
                prod_name,
                category,
                stock,
                min_stock,
                qty_to_order,
                cost_price,
                total_cost,
                unit,
                status,
            ]

            for col_idx, value in enumerate(row_data, start=2):
                cell = ws3.cell(row=row_num, column=col_idx, value=value)
                cell.font = font_body()
                row_fill = fill_data_row(counter)
                cell.fill = row_fill

                if isinstance(value, (int, float)):
                    cell.alignment = align_number()
                    if col_idx in (9, 10):  # Prix, Coût
                        cell.number_format = CURRENCY_FORMAT
                    elif col_idx in (6, 7, 8):
                        cell.number_format = '#,##0'
                else:
                    cell.alignment = align_text()

            # Color status
            status_cell = ws3.cell(row=row_num, column=12)  # col L
            if status == "À commander":
                status_cell.font = Font(name=FONT_NAME, size=11, color=ACCENT_NEGATIVE, bold=HEADER_BOLD)
            else:
                status_cell.font = Font(name=FONT_NAME, size=11, color=ACCENT_POSITIVE)

            ws3.row_dimensions[row_num].height = ROW_HEIGHTS["data"]
            row_num += 1

    detail_last_data = row_num - 1

    # Total row
    ws3.cell(row=row_num, column=2, value="").font = font_subheader()
    ws3.cell(row=row_num, column=3, value="TOTAL").font = font_subheader()
    ws3.cell(row=row_num, column=3).alignment = align_text()

    for col_letter_idx in [7, 8, 10]:  # Qty, (not price), Cost
        col_letter = get_column_letter(col_letter_idx)
        cell = ws3.cell(row=row_num, column=col_letter_idx,
                        value=f"=SUM({col_letter}5:{col_letter}{detail_last_data})")
        cell.alignment = align_number()
        if col_letter_idx == 10:
            cell.number_format = CURRENCY_FORMAT
        else:
            cell.number_format = '#,##0'

    style_total_row(ws3, row_num=row_num, col_start=2, col_end=detail_last_col)

    # Column widths
    for col_idx, (_, width) in enumerate(detail_headers, start=2):
        ws3.column_dimensions[get_column_letter(col_idx)].width = width

    # Conditional formatting on Qty to order column
    qty_range = f"H5:H{detail_last_data}"
    ws3.conditional_formatting.add(qty_range,
        CellIsRule(operator='greaterThan', formula=['0'],
                   fill=CF_NEGATIVE_FILL, font=CF_NEGATIVE_FONT))

    # Auto filter
    ws3.auto_filter.ref = f"B4:L{detail_last_data}"

    ws3.freeze_panes = "B5"
    ws3.page_setup.orientation = 'landscape'

    # ============================================================
    # Save
    # ============================================================
    wb.properties.creator = "Z.ai"
    output_path = "/home/z/my-project/download/Liste_Produits_A_Commander_Par_Fournisseur.xlsx"
    wb.save(output_path)
    print(f"Fichier généré : {output_path}")
    return output_path


if __name__ == "__main__":
    path = build_workbook()
